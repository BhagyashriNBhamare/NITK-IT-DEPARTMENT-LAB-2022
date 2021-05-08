from openpyxl import load_workbook
workbook = load_workbook(filename="4.csv")
print(sheet.cell(row=2, column=2))